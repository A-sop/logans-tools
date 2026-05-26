#!/usr/bin/env python3
"""Export master EÜR year sheet rows to JSON for LDW Books UI (no Zoho API)."""

from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

DEFAULT_MASTER = Path(r'C:\DATA\20_ADMIN\!!_TAX-ADMIN\250111_LDW_EÜR-seit2020-m-Vorlage.xlsx')
DETAIL_START = 9
ZEILE_RE = re.compile(r'^(\d{1,3})')


def zeile_from_label(raw: object) -> str:
    if raw is None:
        return ''
    text = str(raw).strip()
    if ' - ' in text:
        text = text.split(' - ', 1)[0].strip()
    m = ZEILE_RE.match(text)
    return m.group(1) if m else ''


def format_date(raw: object) -> str:
    if raw is None:
        return ''
    if isinstance(raw, datetime):
        return raw.strftime('%d.%m.%Y')
    if isinstance(raw, date):
        return raw.strftime('%d.%m.%Y')
    return str(raw).strip()


def export_sheet(ws, year: int) -> list[dict]:
    rows_out: list[dict] = []
    for row in range(DETAIL_START, ws.max_row + 1):
        datum = ws.cell(row, 2).value
        zeile_raw = ws.cell(row, 3).value
        notiz = ws.cell(row, 4).value
        income = ws.cell(row, 5).value
        expense = ws.cell(row, 8).value

        if str(datum).strip().lower() == 'datum':
            continue

        inc = float(income) if isinstance(income, (int, float)) and income > 0 else 0.0
        exp = float(expense) if isinstance(expense, (int, float)) and expense > 0 else 0.0
        if inc <= 0 and exp <= 0:
            continue

        payee = str(notiz or '').strip() or '—'
        zeile_label = str(zeile_raw or '').strip()
        zeile = zeile_from_label(zeile_raw)
        if inc > 0:
            direction = 'in'
            amount = inc
        else:
            direction = 'out'
            amount = exp

        rows_out.append(
            {
                'date': format_date(datum),
                'payee': payee,
                'direction': direction,
                'amount': round(amount, 2),
                'eurZeile': zeile,
                'eurLabel': zeile_label,
            }
        )
    return rows_out


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument('--master', type=Path, default=DEFAULT_MASTER)
    parser.add_argument('--year', type=int, action='append')
    parser.add_argument(
        '--dest',
        type=Path,
        default=Path(__file__).resolve().parent.parent / 'src' / 'data' / 'euer',
    )
    args = parser.parse_args()
    years = args.year or list(range(2020, 2027))
    if not args.master.is_file():
        raise SystemExit(f'Master workbook not found: {args.master}')

    wb = openpyxl.load_workbook(args.master, read_only=True, data_only=True)
    for year in years:
        sheet = f'{year}_EÜR'
        out = args.dest / f'transactions-workbook-{year}.json'
        if sheet not in wb.sheetnames:
            print(f'{year}: sheet missing, skip')
            continue
        rows_out = export_sheet(wb[sheet], year)
        payload = {
            'year': year,
            'exported': date.today().isoformat(),
            'source': str(args.master),
            'rows': rows_out,
        }
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
        print(f'{year}: {len(rows_out)} rows -> {out.name}')
    wb.close()


if __name__ == '__main__':
    main()
